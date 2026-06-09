#!/usr/bin/env python3
"""Excel MCP Server — cross-platform Excel file automation via openpyxl.

Communicates over stdio using the MCP JSON-RPC protocol.
"""

import json
import sys
import traceback
from pathlib import Path

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter, column_index_from_string
except ImportError:
    print("Missing openpyxl. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def read_message():
    buf = b""
    while True:
        chunk = sys.stdin.buffer.read1(4096)
        if not chunk:
            return None
        buf += chunk
        idx = buf.find(b"\r\n\r\n")
        if idx == -1:
            continue
        header = buf[:idx].decode("ascii")
        match = __import__("re").search(r"Content-Length:\s*(\d+)", header)
        if not match:
            buf = buf[idx + 4:]
            continue
        length = int(match.group(1))
        body_start = idx + 4
        if len(buf) < body_start + length:
            continue
        body = buf[body_start:body_start + length]
        buf = buf[body_start + length:]
        return json.loads(body.decode("utf-8"))


def write_message(msg):
    body = json.dumps(msg, ensure_ascii=False)
    header = f"Content-Length: {len(body.encode('utf-8'))}\r\n\r\n"
    sys.stdout.buffer.write(header.encode("ascii") + body.encode("utf-8"))
    sys.stdout.buffer.flush()


def send_result(id, result):
    write_message({"jsonrpc": "2.0", "id": id, "result": result})


def send_error(id, code, message):
    write_message({"jsonrpc": "2.0", "id": id, "error": {"code": code, "message": message}})


# ---------------------------------------------------------------------------
# Tool implementations
# ---------------------------------------------------------------------------

WORKBOOKS: dict[str, Workbook] = {}


def _resolve_path(path: str) -> str:
    return str(Path(path).expanduser().resolve())


TOOLS = [
    {
        "name": "excel_open",
        "description": "Open an existing Excel workbook and keep it in memory",
        "inputSchema": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Path to the .xlsx file"}
            },
            "required": ["path"]
        }
    },
    {
        "name": "excel_create",
        "description": "Create a new Excel workbook (optionally save to path)",
        "inputSchema": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Where to save the new workbook"},
                "sheet_name": {"type": "string", "description": "Name for the initial sheet (default: Sheet1)"}
            },
            "required": ["path"]
        }
    },
    {
        "name": "excel_list_sheets",
        "description": "List all sheet names in an open workbook",
        "inputSchema": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Path of the open workbook"}
            },
            "required": ["path"]
        }
    },
    {
        "name": "excel_read",
        "description": "Read cell values from a sheet. Returns a range or single cell as JSON",
        "inputSchema": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Path of the open workbook"},
                "sheet": {"type": "string", "description": "Sheet name"},
                "range": {"type": "string", "description": "Cell range like A1:C5 or A1 (default: used range)"}
            },
            "required": ["path", "sheet"]
        }
    },
    {
        "name": "excel_write",
        "description": "Write values to a sheet. Provide cell references and values",
        "inputSchema": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Path of the open workbook"},
                "sheet": {"type": "string", "description": "Sheet name"},
                "data": {
                    "type": "object",
                    "description": "Mapping of cell references to values, e.g. {\"A1\": \"Name\", \"B1\": \"Age\"}"
                }
            },
            "required": ["path", "sheet", "data"]
        }
    },
    {
        "name": "excel_write_range",
        "description": "Write a 2D array of values starting at a given cell",
        "inputSchema": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Path of the open workbook"},
                "sheet": {"type": "string", "description": "Sheet name"},
                "start_cell": {"type": "string", "description": "Top-left cell, e.g. A1"},
                "values": {
                    "type": "array",
                    "items": {"type": "array"},
                    "description": "2D array of row values"
                }
            },
            "required": ["path", "sheet", "start_cell", "values"]
        }
    },
    {
        "name": "excel_save",
        "description": "Save an open workbook to disk",
        "inputSchema": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Path of the open workbook"}
            },
            "required": ["path"]
        }
    },
    {
        "name": "excel_close",
        "description": "Close a workbook (removes it from memory without saving)",
        "inputSchema": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Path of the open workbook"}
            },
            "required": ["path"]
        }
    },
    {
        "name": "excel_create_sheet",
        "description": "Create a new sheet in an open workbook",
        "inputSchema": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Path of the open workbook"},
                "name": {"type": "string", "description": "Name for the new sheet"}
            },
            "required": ["path", "name"]
        }
    },
    {
        "name": "excel_delete_sheet",
        "description": "Delete a sheet from an open workbook",
        "inputSchema": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Path of the open workbook"},
                "name": {"type": "string", "description": "Sheet name to delete"}
            },
            "required": ["path", "name"]
        }
    },
]

TOOL_HANDLERS = {}


def tool(name):
    def decorator(fn):
        TOOL_HANDLERS[name] = fn
        return fn
    return decorator


@tool("excel_open")
def handle_open(args):
    path = _resolve_path(args["path"])
    if path in WORKBOOKS:
        return {"message": f"Workbook already open: {path}"}
    wb = load_workbook(path, data_only=True)
    WORKBOOKS[path] = wb
    sheets = wb.sheetnames
    return {"message": f"Opened {path}", "sheets": sheets}


@tool("excel_create")
def handle_create(args):
    path = _resolve_path(args["path"])
    sheet_name = args.get("sheet_name", "Sheet1")
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    wb.save(path)
    WORKBOOKS[path] = wb
    return {"message": f"Created {path}", "sheets": wb.sheetnames}


@tool("excel_list_sheets")
def handle_list_sheets(args):
    path = _resolve_path(args["path"])
    wb = _get_wb(path)
    return {"sheets": wb.sheetnames}


@tool("excel_read")
def handle_read(args):
    path = _resolve_path(args["path"])
    wb = _get_wb(path)
    sheet_name = args["sheet"]
    if sheet_name not in wb.sheetnames:
        return {"error": f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"}
    ws = wb[sheet_name]

    if "range" in args:
        cells = ws[args["range"]]
        if isinstance(cells, tuple):
            rows = []
            for row in cells:
                rows.append([cell.value for cell in (row if isinstance(row, (list, tuple)) else [row])])
            return {"data": rows}
        else:
            return {"data": [[cells.value]]}
    else:
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True):
            rows.append(list(row))
        return {"data": rows, "dimensions": f"{ws.dimensions}"}


@tool("excel_write")
def handle_write(args):
    path = _resolve_path(args["path"])
    wb = _get_wb(path)
    sheet_name = args["sheet"]
    if sheet_name not in wb.sheetnames:
        return {"error": f"Sheet '{sheet_name}' not found"}
    ws = wb[sheet_name]
    for cell_ref, value in args["data"].items():
        ws[cell_ref] = value
    wb.save(path)
    return {"message": f"Wrote {len(args['data'])} cell(s)", "updated_cells": list(args["data"].keys())}


@tool("excel_write_range")
def handle_write_range(args):
    path = _resolve_path(args["path"])
    wb = _get_wb(path)
    sheet_name = args["sheet"]
    if sheet_name not in wb.sheetnames:
        return {"error": f"Sheet '{sheet_name}' not found"}
    ws = wb[sheet_name]

    start_col_letter = "".join(c for c in args["start_cell"] if c.isalpha()).upper()
    start_row = int("".join(c for c in args["start_cell"] if c.isdigit()))
    start_col = column_index_from_string(start_col_letter)

    for i, row in enumerate(args["values"]):
        for j, value in enumerate(row):
            cell = ws.cell(row=start_row + i, column=start_col + j)
            cell.value = value

    written = sum(len(r) for r in args["values"])
    wb.save(path)
    return {"message": f"Wrote {written} values starting at {args['start_cell']}"}


@tool("excel_save")
def handle_save(args):
    path = _resolve_path(args["path"])
    wb = _get_wb(path)
    wb.save(path)
    return {"message": f"Saved {path}"}


@tool("excel_close")
def handle_close(args):
    path = _resolve_path(args["path"])
    if path in WORKBOOKS:
        WORKBOOKS[path].close()
        del WORKBOOKS[path]
        return {"message": f"Closed {path}"}
    return {"message": f"Workbook not open: {path}"}


@tool("excel_create_sheet")
def handle_create_sheet(args):
    path = _resolve_path(args["path"])
    wb = _get_wb(path)
    wb.create_sheet(title=args["name"])
    wb.save(path)
    return {"message": f"Created sheet '{args['name']}'", "sheets": wb.sheetnames}


@tool("excel_delete_sheet")
def handle_delete_sheet(args):
    path = _resolve_path(args["path"])
    wb = _get_wb(path)
    name = args["name"]
    if name not in wb.sheetnames:
        return {"error": f"Sheet '{name}' not found"}
    std = wb[name]
    wb.remove(std)
    wb.save(path)
    return {"message": f"Deleted sheet '{name}'", "sheets": wb.sheetnames}


def _get_wb(path: str) -> Workbook:
    path = _resolve_path(path)
    if path not in WORKBOOKS:
        raise ValueError(f"Workbook not open: {path}. Use excel_open first.")
    return WORKBOOKS[path]


# ---------------------------------------------------------------------------
# Main loop
# ---------------------------------------------------------------------------

def main():
    while True:
        msg = read_message()
        if msg is None:
            break

        method = msg.get("method")
        msg_id = msg.get("id")
        params = msg.get("params", {})

        if method == "initialize":
            send_result(msg_id, {
                "protocolVersion": "0.1",
                "capabilities": {"tools": {}},
                "serverInfo": {"name": "excel-mcp", "version": "0.1.0"}
            })

        elif method == "initialized":
            continue

        elif method == "tools/list":
            send_result(msg_id, {"tools": TOOLS})

        elif method == "tools/call":
            name = params.get("name")
            arguments = params.get("arguments", {})

            handler = TOOL_HANDLERS.get(name)
            if not handler:
                send_error(msg_id, -32601, f"Tool not found: {name}")
                continue

            try:
                result = handler(arguments)
                if isinstance(result, dict) and "error" in result:
                    send_result(msg_id, {
                        "content": [{"type": "text", "text": result["error"]}],
                        "isError": True
                    })
                else:
                    send_result(msg_id, {
                        "content": [{"type": "text", "text": json.dumps(result, indent=2)}]
                    })
            except Exception as e:
                send_result(msg_id, {
                    "content": [{"type": "text", "text": f"{type(e).__name__}: {e}\n{traceback.format_exc()}"}],
                    "isError": True
                })

        elif method == "shutdown":
            send_result(msg_id, {})
            break

        else:
            send_error(msg_id, -32601, f"Method not found: {method}")

    for wb in WORKBOOKS.values():
        try:
            wb.close()
        except Exception:
            pass


if __name__ == "__main__":
    main()
